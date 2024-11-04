import requests
import pytest
import openpyxl
from src.constants.api_constants import url_create_token
from src.helpers.utils import common_headers_json


def read_credentials_from_excel(file_path):
    credentials = []
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active
    for rows in sheet.iter_rows(min_row=2, values_only=True):
        username, password = rows
        credentials.append({"username": username,"password": password})
    return credentials


def make_request_auth(username, password):
    payload = {
        "username": username,
        "password": password
    }
    response = requests.post(url=url_create_token(), headers=common_headers_json(), json=payload)
    return response


def test_post_create_token():
    file_path = "/Users/Dhivya/PycharmProjects/pythonAPIProject/tests/datadriventesting/input.xlsx"
    credentials = read_credentials_from_excel(file_path)

    for user_cred in credentials:
        username = user_cred["username"]
        password = user_cred["password"]
        print(username, password)
        response = make_request_auth(username, password)
        print(response)
# wb=openpyxl.load_workbook("data.xlsx")
        assert response.status_code==200
        # if response.json()["token"]==True:
        #     print("Pass")
        # else:
        #     print("Fail")
